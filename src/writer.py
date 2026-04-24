"""
Results writer.
Writes simulation outputs back to the Portfolio State workbook output tabs.
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def write_portfolio_outputs(
    state_path: str,
    summary: dict,
    corr_index: float,
    run_label: str = "RUN-AUTO",
    notes: str = "",
):
    """
    Write portfolio-level simulation results to SIM_OUTPUTS_PORTFOLIO tab.
    Appends a new row; does not overwrite existing rows.
    """
    wb = load_workbook(state_path)
    ws = wb["SIM_OUTPUTS_PORTFOLIO"]

    # Find first empty row (after header)
    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(2, 1).value is None:
        next_row = 2

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    values = [
        run_label,                                          # Run_ID
        now,                                                # AsOf_Date
        round(summary.get("P_ThreePlus_Exits", 0), 4),     # P_Exits_GE3
        round(summary.get("P_Exits_GE2", 0), 4),           # P_Exits_GE2
        round(summary.get("P_Exits_LE1", 0), 4),           # P_Exits_LE1
        round(summary.get("Median_Annual_IRR", 0), 4),     # Median_IRR
        round(summary.get("IRR_P10", 0), 4),               # IRR_P10
        round(summary.get("Median_MOIC", 0), 4),           # Median_MOIC
        round(summary.get("MOIC_P10", 0), 4),              # MOIC_P10
        round(summary.get("Median_First_Dist_Month", 0), 1),  # Median_TimeToFirstDist
        round(corr_index, 4),                               # CorrIndex
        notes,                                               # Notes
    ]

    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(state_path)
    print(f"  Written portfolio outputs to row {next_row}")


def write_results_to_console(summary: dict, envelope_result: dict, signal: dict):
    """Print formatted results to console."""
    print("\n" + "=" * 60)
    print(f"  PORTFOLIO QUALITY SIGNAL: {signal['label']}")
    print(f"  {signal['message']}")
    print("=" * 60)

    print(f"\n  Simulations:        {summary['Num_Sims']:,}")
    print(f"  Median MOIC:        {summary['Median_MOIC']:.2f}x")
    print(f"  MOIC P10:           {summary['MOIC_P10']:.2f}x")
    print(f"  Median Annual IRR:  {summary['Median_Annual_IRR']:.1%}")
    print(f"  IRR P10:            {summary['IRR_P10']:.1%}")
    print(f"  P(0 exits):         {summary['P_Zero_Exits']:.1%}")
    print(f"  P(1 exit):          {summary['P_One_Exit']:.1%}")
    print(f"  P(2 exits):         {summary['P_Two_Exits']:.1%}")
    print(f"  P(3+ exits):        {summary['P_ThreePlus_Exits']:.1%}")

    if summary.get("Median_First_Dist_Month") and pd.notna(summary["Median_First_Dist_Month"]):
        print(f"  Median 1st Dist:    Month {summary['Median_First_Dist_Month']:.0f}")

    print(f"\n  Envelope Checks:")
    for name, check in envelope_result["checks"].items():
        status = "PASS" if check["pass"] else "FAIL"
        actual = check["actual"]
        threshold = check["threshold"]
        if isinstance(actual, float):
            print(f"    {name:20s}  {status}  (actual={actual:.4f}, threshold={threshold})")
        else:
            print(f"    {name:20s}  {status}  (actual={actual}, threshold={threshold})")

    if envelope_result["failed_gates"]:
        print(f"\n  FAILED GATES: {', '.join(envelope_result['failed_gates'])}")
    else:
        print(f"\n  ALL GATES PASSED")
    print()


def write_contribution_outputs(
    state_path: str,
    candidate_id: str,
    contribution: dict,
    run_label: str = "RUN-AUTO",
    notes: str = "",
):
    """Write contribution results to SIM_OUTPUTS_ASSET_CONTRIBUTION tab."""
    wb = load_workbook(state_path)
    ws = wb["SIM_OUTPUTS_ASSET_CONTRIBUTION"]

    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(2, 1).value is None:
        next_row = 2

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    values = [
        run_label,                          # Run_ID
        now,                                # AsOf_Date
        candidate_id,                       # Candidate_Asset_ID
        contribution["EDC"],                # EDC
        contribution["IRC"],                # IRC
        contribution["DPC"],                # DPC
        contribution.get("LAC_Months"),     # LAC_Months
        contribution["CDC"],                # CDC
        contribution["EDC_Low"],
        contribution["EDC_Base"],
        contribution["EDC_High"],
        contribution["IRC_Low"],
        contribution["IRC_Base"],
        contribution["IRC_High"],
        contribution["DPC_Low"],
        contribution["DPC_Base"],
        contribution["DPC_High"],
        contribution.get("LAC_Low"),
        contribution.get("LAC_Base"),
        contribution.get("LAC_High"),
        contribution["CDC_Low"],
        contribution["CDC_Base"],
        contribution["CDC_High"],
        notes,
    ]

    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(state_path)
    print(f"  Written contribution outputs for {candidate_id} to row {next_row}")


def write_mec_outputs(
    state_path: str,
    candidate_id: str,
    mec_result: dict,
    run_label: str = "RUN-AUTO",
    notes: str = "",
):
    """Write MEC results to MEC_RESULTS tab."""
    wb = load_workbook(state_path)
    ws = wb["MEC_RESULTS"]

    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(2, 1).value is None:
        next_row = 2

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    values = [
        run_label,
        now,
        candidate_id,
        mec_result.get("MEC_AcqCash", 0),
        mec_result.get("MEC_Equity", 0),
        mec_result.get("MEC_PassThrough", 0),
        mec_result.get("MEC_DeferredCash", 0),
        mec_result.get("first_failed_constraint", ""),
        notes,
    ]

    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(state_path)
    print(f"  Written MEC results for {candidate_id} to row {next_row}")
