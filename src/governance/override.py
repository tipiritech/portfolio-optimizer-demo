"""
Executive Override System.
Allows Board of Directors to override specific governance gate failures.

Rules from Governance Doctrine:
  - Requires Board supermajority (≥80%) vote
  - Must specify which gates are being overridden
  - Must include written justification
  - Override is time-limited (expires after defined period)
  - Full audit trail required (Run_ID, date, vote, justification)
  - Override does NOT change locked parameters — it accepts the risk
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook


# Gates that CAN be overridden by Board vote
OVERRIDABLE_GATES = {
    "Median_IRR",
    "IRR_P10",
    "P_Exits_GE3",
    "P_Exits_LE1",
    "Corr_Index",
    "Weighted_Time",
    "Capital_Concentration",
    "Capital_Pause_Trigger",
    "EDC_Positive",
    "EDC_Above_Stop",
    "Catastrophic",
    "Correlation",
    "Tight_Stress",
    "Hard_Cap",
}

# Gates that can NEVER be overridden — hard structural constraints
NON_OVERRIDABLE = {
    "Max_Duration",  # 36-month hard stop is non-negotiable
}

# Default override expiration: 90 days
DEFAULT_EXPIRY_DAYS = 90


def create_override(
    failed_gates: list,
    board_vote: str,
    justification: str,
    approved_by: str,
    asset_id: str | None = None,
    run_id: str | None = None,
    expiry_days: int = DEFAULT_EXPIRY_DAYS,
    conditions: str = "",
    risk_accepted: str = "",
) -> dict:
    """
    Create an executive override record.

    Args:
        failed_gates: list of gate names being overridden
        board_vote: vote tally (e.g., "4/5", "5/5")
        justification: written reason for override
        approved_by: name/role of approver
        asset_id: optional — if override is for a specific asset admission
        run_id: Monte Carlo run ID that produced the failure
        expiry_days: days until override expires (default 90)
        conditions: any conditions attached to the override
        risk_accepted: explicit statement of risk being accepted

    Returns:
        dict: override record with validation status
    """
    now = datetime.now()

    # Validate vote meets supermajority
    vote_valid = _validate_vote(board_vote)

    # Check all gates are overridable
    non_overridable_attempted = [g for g in failed_gates if g in NON_OVERRIDABLE]
    unknown_gates = [g for g in failed_gates if g not in OVERRIDABLE_GATES and g not in NON_OVERRIDABLE]

    override = {
        "Override_ID": f"OVR-{now.strftime('%Y%m%d-%H%M%S')}",
        "Created_Date": now.strftime("%Y-%m-%d %H:%M:%S"),
        "Expiry_Date": (now + timedelta(days=expiry_days)).strftime("%Y-%m-%d"),
        "Asset_ID": asset_id,
        "Run_ID": run_id,
        "Failed_Gates": ", ".join(failed_gates),
        "Board_Vote": board_vote,
        "Vote_Valid": vote_valid,
        "Approved_By": approved_by,
        "Justification": justification,
        "Conditions": conditions,
        "Risk_Accepted": risk_accepted,
        "Status": "PENDING",
        "Non_Overridable_Attempted": ", ".join(non_overridable_attempted) if non_overridable_attempted else None,
        "Unknown_Gates": ", ".join(unknown_gates) if unknown_gates else None,
    }

    # Determine if override is valid
    if non_overridable_attempted:
        override["Status"] = "REJECTED"
        override["Rejection_Reason"] = f"Cannot override: {', '.join(non_overridable_attempted)}"
    elif not vote_valid:
        override["Status"] = "REJECTED"
        override["Rejection_Reason"] = f"Vote '{board_vote}' does not meet supermajority (≥80%)"
    elif not justification.strip():
        override["Status"] = "REJECTED"
        override["Rejection_Reason"] = "Justification is required"
    else:
        override["Status"] = "APPROVED"
        override["Rejection_Reason"] = None

    return override


def _validate_vote(vote_str: str) -> bool:
    """
    Check if vote meets ≥80% supermajority.
    Accepts formats: "4/5", "5/5", "80%", etc.
    """
    vote_str = vote_str.strip()

    # Format: "X/Y"
    if "/" in vote_str:
        try:
            parts = vote_str.split("/")
            yes = float(parts[0])
            total = float(parts[1])
            if total <= 0:
                return False
            return (yes / total) >= 0.80
        except (ValueError, IndexError):
            return False

    # Format: "80%" or "0.80"
    try:
        val = float(vote_str.replace("%", ""))
        if val > 1:
            val = val / 100.0
        return val >= 0.80
    except ValueError:
        return False


def apply_override(
    envelope_result: dict,
    override: dict,
) -> dict:
    """
    Apply an approved override to envelope check results.

    Does NOT change the actual metrics — just marks overridden gates
    so downstream logic can proceed while maintaining the audit trail.

    Returns modified envelope_result with override annotations.
    """
    if override["Status"] != "APPROVED":
        return envelope_result

    overridden_gates = [g.strip() for g in override["Failed_Gates"].split(",")]

    modified = {
        "checks": {},
        "all_pass": True,
        "failed_gates": [],
        "overrides_applied": [],
    }

    for gate_name, check in envelope_result["checks"].items():
        new_check = check.copy()

        if not check["pass"] and gate_name in overridden_gates:
            new_check["overridden"] = True
            new_check["override_id"] = override["Override_ID"]
            new_check["original_pass"] = False
            # Gate is considered "passed with override"
            new_check["pass"] = True
            modified["overrides_applied"].append(gate_name)
        else:
            new_check["overridden"] = False

        modified["checks"][gate_name] = new_check

    # Recalculate overall pass (with overrides applied)
    modified["all_pass"] = all(c["pass"] for c in modified["checks"].values())
    modified["failed_gates"] = [k for k, v in modified["checks"].items() if not v["pass"]]

    return modified


def apply_override_to_admission(
    gate_result: dict,
    override: dict,
) -> dict:
    """Apply an approved override to admission gate results."""
    if override["Status"] != "APPROVED":
        return gate_result

    overridden_gates = [g.strip() for g in override["Failed_Gates"].split(",")]

    modified = {
        "gates": {},
        "all_pass": True,
        "failed_gates": [],
        "requires_override": False,
        "override_type": None,
        "overrides_applied": [],
    }

    for gate_name, gate in gate_result["gates"].items():
        new_gate = gate.copy()
        if not gate["pass"] and gate_name in overridden_gates:
            new_gate["overridden"] = True
            new_gate["override_id"] = override["Override_ID"]
            new_gate["original_pass"] = False
            new_gate["pass"] = True
            modified["overrides_applied"].append(gate_name)
        else:
            new_gate["overridden"] = False
        modified["gates"][gate_name] = new_gate

    modified["all_pass"] = all(g["pass"] for g in modified["gates"].values())
    modified["failed_gates"] = [k for k, v in modified["gates"].items() if not v["pass"]]
    modified["requires_override"] = not modified["all_pass"]

    return modified


def check_override_expiry(override: dict) -> bool:
    """Check if an override has expired."""
    expiry = datetime.strptime(override["Expiry_Date"], "%Y-%m-%d")
    return datetime.now() > expiry


def write_override_to_excel(state_path: str, override: dict):
    """Write override record to OVERRIDE_LOG tab in Portfolio State workbook."""
    wb = load_workbook(state_path)
    ws = wb["OVERRIDE_LOG"]

    next_row = ws.max_row + 1
    if next_row == 2 and ws.cell(2, 1).value is None:
        next_row = 2

    values = [
        override["Created_Date"],
        override.get("Asset_ID", ""),
        override["Failed_Gates"],
        override["Board_Vote"],
        override["Justification"],
        f"Status={override['Status']}; "
        f"Override_ID={override['Override_ID']}; "
        f"Expiry={override['Expiry_Date']}; "
        f"Conditions={override.get('Conditions', '')}; "
        f"Risk={override.get('Risk_Accepted', '')}",
    ]

    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=next_row, column=col_idx, value=val)

    wb.save(state_path)
    print(f"  Override {override['Override_ID']} written to OVERRIDE_LOG row {next_row}")


def format_override_report(override: dict) -> str:
    """Format override as a text report."""
    lines = []
    lines.append("=" * 60)
    lines.append(f"  EXECUTIVE OVERRIDE: {override['Override_ID']}")
    lines.append("=" * 60)
    lines.append(f"  Status:          {override['Status']}")
    lines.append(f"  Date:            {override['Created_Date']}")
    lines.append(f"  Expiry:          {override['Expiry_Date']}")
    lines.append(f"  Asset:           {override.get('Asset_ID', 'Portfolio-level')}")
    lines.append(f"  Run ID:          {override.get('Run_ID', 'N/A')}")
    lines.append(f"  Failed Gates:    {override['Failed_Gates']}")
    lines.append(f"  Board Vote:      {override['Board_Vote']} ({'VALID' if override['Vote_Valid'] else 'INVALID'})")
    lines.append(f"  Approved By:     {override['Approved_By']}")
    lines.append(f"  Justification:   {override['Justification']}")

    if override.get("Conditions"):
        lines.append(f"  Conditions:      {override['Conditions']}")
    if override.get("Risk_Accepted"):
        lines.append(f"  Risk Accepted:   {override['Risk_Accepted']}")

    if override["Status"] == "REJECTED":
        lines.append(f"  REJECTION:       {override.get('Rejection_Reason', 'Unknown')}")

    if override.get("Non_Overridable_Attempted"):
        lines.append(f"  NON-OVERRIDABLE: {override['Non_Overridable_Attempted']}")

    lines.append("")
    return "\n".join(lines)
